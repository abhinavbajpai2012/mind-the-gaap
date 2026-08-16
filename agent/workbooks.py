"""Fill the challenge template workbooks from a finished run.

The templates in `challenge/templates/` are the ground truth for layout: the
`Summary` sheet, the metric labels, units and the fiscal-period column must all
survive unchanged, so each workbook starts as a copy of its template and only
the yellow forecast cells are written. Percentage metrics are already carried
in percentage points end to end, which is what the template asks for.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


def write_workbooks(run_result, root: Path) -> list[Path]:
    """Write one workbook per company in the run; return the paths written.

    A company whose three metrics are not all filled is skipped with a
    ValueError rather than uploaded half-empty — the checker would reject it
    anyway, and a silent blank cell is worse than a loud failure.
    """
    definitions = json.loads(
        (root / "challenge" / "companies.json").read_text(encoding="utf-8"))
    out_dir = root / "submission"
    out_dir.mkdir(exist_ok=True)

    by_key = {(r.ticker, r.metric): r for r in run_result.results}
    written: list[Path] = []
    problems: list[str] = []

    for company in definitions["companies"]:
        ticker = company["ticker"].split(":")[-1]   # "LSE:HAS" -> "HAS"
        if ticker not in run_result.companies:
            continue

        template = root / "challenge" / "templates" / company["outputFile"]
        book = load_workbook(template)
        sheet = book["Summary"]

        header_row = next(
            (row for row in range(1, 31)
             if str(sheet.cell(row, 1).value or "").strip() == "Metric"
             and str(sheet.cell(row, 2).value or "").strip() == "Units"
             and str(sheet.cell(row, 3).value or "").strip() == company["period"]),
            None)
        if header_row is None:
            problems.append(f"{company['outputFile']}: header row not found")
            continue

        missing = []
        for offset, metric in enumerate(company["metrics"], start=1):
            result = by_key.get((ticker, metric["label"]))
            if result is None or result.value is None:
                missing.append(
                    f"{company['outputFile']}: no forecast for {metric['label']}"
                    + (f" ({result.error})" if result and result.error else ""))
                continue
            sheet.cell(header_row + offset, 3).value = round(result.value, 4)

        if missing:
            problems += missing
            continue
        target = out_dir / company["outputFile"]
        book.save(target)
        written.append(target)

    if problems:
        raise ValueError("; ".join(problems))
    return written
